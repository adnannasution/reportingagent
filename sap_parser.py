"""
sap_parser.py — Parse Excel SAP (IW29 / IW39) ke format DB
Support: IW29 (Notification), IW39 PT02/PT03/PT05/PT08 (Work Orders)
"""

import openpyxl
from datetime import datetime, date


def to_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    return None

def to_str(val):
    if val is None:
        return None
    return str(val).strip() or None

def to_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except:
        return None


def detect_file_type(header: tuple) -> str:
    """Deteksi tipe file dari header kolom."""
    cols = [str(c).strip() if c else '' for c in header]
    if 'Notifictn type' in cols or 'Notification' in cols and 'Order Type' not in cols:
        return 'notification'
    if 'Order Type' in cols and 'Bas. start date' in cols:
        return 'work_order'
    if 'BOM category' in cols and 'Component' in cols:
        return 'bom'
    return 'unknown'


def parse_notification_file(filepath: str, batch_id: str) -> list:
    """Parse IW29 Excel → list of tuples untuk insert."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    header = [str(c).strip() if c else '' for c in rows[0]]

    def col(name):
        try:
            return header.index(name)
        except ValueError:
            return None

    results = []
    for r in rows[1:]:
        if not any(c is not None for c in r):
            continue

        has_long = False
        lt_idx = col('Long text')
        if lt_idx is not None and r[lt_idx]:
            has_long = str(r[lt_idx]).strip() in ('X', 'x', '1', 'True')

        results.append((
            to_str(r[col('Notifictn type')] if col('Notifictn type') is not None else None),
            to_date(r[col('Notif.date')] if col('Notif.date') is not None else None),
            to_str(r[col('Notification')] if col('Notification') is not None else None),
            to_str(r[col('System status')] if col('System status') is not None else None),
            to_date(r[col('Req. start')] if col('Req. start') is not None else None),
            to_date(r[col('Required End')] if col('Required End') is not None else None),
            to_str(r[col('Main WorkCtr')] if col('Main WorkCtr') is not None else None),
            to_str(r[col('Planner group')] if col('Planner group') is not None else None),
            to_str(r[col('Description')] if col('Description') is not None else None),
            to_str(r[col('Order')] if col('Order') is not None else None),
            to_str(r[col('Location')] if col('Location') is not None else None),
            to_str(r[col('Functional Loc.')] if col('Functional Loc.') is not None else None),
            to_str(r[col('Equipment')] if col('Equipment') is not None else None),
            to_str(r[col('Criticallity')] if col('Criticallity') is not None else None),
            to_str(r[col('MaintPlant')] if col('MaintPlant') is not None else None),
            has_long,
            batch_id,
        ))
    return results


def parse_work_order_file(filepath: str, batch_id: str) -> list:
    """Parse IW39 Excel → list of tuples untuk insert."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    header = [str(c).strip() if c else '' for c in rows[0]]

    def col(name):
        try:
            return header.index(name)
        except ValueError:
            return None

    results = []
    for r in rows[1:]:
        if not any(c is not None for c in r):
            continue

        # Ambil status utama dari system_status
        sys_st = to_str(r[col('System status')] if col('System status') is not None else None)
        user_st_raw = r[col('User status')] if col('User status') is not None else None
        user_st = to_str(user_st_raw)

        results.append((
            to_str(r[col('Plant')] if col('Plant') is not None else None),
            to_date(r[col('Created on')] if col('Created on') is not None else None),
            to_date(r[col('Changed on')] if col('Changed on') is not None else None),
            to_date(r[col('Bas. start date')] if col('Bas. start date') is not None else None),
            to_date(r[col('Basic fin. date')] if col('Basic fin. date') is not None else None),
            to_str(r[col('Notification')] if col('Notification') is not None else None),
            to_str(r[col('Order')] if col('Order') is not None else None),
            to_str(r[col('Superior order')] if col('Superior order') is not None else None),
            to_str(r[col('Description')] if col('Description') is not None else None),
            to_str(r[col('Functional Loc.')] if col('Functional Loc.') is not None else None),
            to_str(r[col('Location')] if col('Location') is not None else None),
            to_str(r[col('Equipment')] if col('Equipment') is not None else None),
            to_str(r[col('Criticallity')] if col('Criticallity') is not None else None),
            user_st,
            sys_st,
            to_str(r[col('Planner group')] if col('Planner group') is not None else None),
            to_float(r[col('TotalPlnndCosts')] if col('TotalPlnndCosts') is not None else None),
            to_float(r[col('Total act.costs')] if col('Total act.costs') is not None else None),
            to_str(r[col('Main WorkCtr')] if col('Main WorkCtr') is not None else None),
            to_str(r[col('PO number')] if col('PO number') is not None else None),
            to_date(r[col('Actual Finish')] if col('Actual Finish') is not None else None),
            to_date(r[col('Actual release')] if col('Actual release') is not None else None),
            to_str(r[col('Order Type')] if col('Order Type') is not None else None),
            to_str(r[col('Priority')] if col('Priority') is not None else None),
            to_str(r[col('MaintActivType')] if col('MaintActivType') is not None else None),
            batch_id,
        ))
    return results


def parse_bom_file(filepath: str, batch_id: str) -> list:
    """Parse BOM Excel → list of tuples untuk insert."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    header = [str(c).strip() if c else '' for c in rows[0]]

    def col(name):
        try:
            return header.index(name)
        except ValueError:
            return None

    # Kolom Description muncul dua kali (equipment & component), ambil index keduanya
    desc_indices = [i for i, h in enumerate(header) if h == 'Description']
    eq_desc_idx   = desc_indices[0] if len(desc_indices) > 0 else None
    comp_desc_idx = desc_indices[1] if len(desc_indices) > 1 else None

    # Item text juga bisa muncul dua kali
    itext_indices = [i for i, h in enumerate(header) if h.lower() == 'item text']
    item_text_idx = itext_indices[0] if itext_indices else None

    results = []
    for r in rows[1:]:
        if not any(c is not None for c in r):
            continue
        results.append((
            to_str(r[col('Equipment')] if col('Equipment') is not None else None),
            to_str(r[eq_desc_idx] if eq_desc_idx is not None else None),
            to_str(r[col('Material')] if col('Material') is not None else None),
            to_str(r[col('Plant')] if col('Plant') is not None else None),
            to_str(r[col('Usage')] if col('Usage') is not None else None),
            to_str(r[col('Item node')] if col('Item node') is not None else None),
            to_str(r[col('BOM category')] if col('BOM category') is not None else None),
            to_str(r[col('EquipCategory')] if col('EquipCategory') is not None else None),
            to_str(r[col('Criticallity')] if col('Criticallity') is not None else None),
            to_str(r[col('Alternative')] if col('Alternative') is not None else None),
            to_str(r[col('Component')] if col('Component') is not None else None),
            to_str(r[comp_desc_idx] if comp_desc_idx is not None else None),
            to_str(r[col('Mfr Part Number')] if col('Mfr Part Number') is not None else None),
            to_str(r[col('Old matl number')] if col('Old matl number') is not None else None),
            to_str(r[col('Material Type')] if col('Material Type') is not None else None),
            to_str(r[col('Item')] if col('Item') is not None else None),
            to_str(r[col('Item Category')] if col('Item Category') is not None else None),
            to_float(r[col('Quantity')] if col('Quantity') is not None else None),
            to_str(r[col('Component unit')] if col('Component unit') is not None else None),
            to_str(r[col('Assembly')] if col('Assembly') is not None else None),
            to_str(r[col('Sort String')] if col('Sort String') is not None else None),
            to_str(r[col('Spare part ID')] if col('Spare part ID') is not None else None),
            to_str(r[item_text_idx] if item_text_idx is not None else None),
            to_str(r[col('Cost element')] if col('Cost element') is not None else None),
            to_str(r[col('Purch. Group')] if col('Purch. Group') is not None else None),
            to_date(r[col('Valid From')] if col('Valid From') is not None else None),
            to_date(r[col('Valid To')] if col('Valid To') is not None else None),
            batch_id,
        ))
    return results


def parse_file(filepath: str, batch_id: str) -> dict:
    """Auto-detect dan parse file SAP Excel."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    first_row = next(ws.iter_rows(max_row=1, values_only=True))
    wb.close()

    file_type = detect_file_type(first_row)

    if file_type == 'notification':
        rows = parse_notification_file(filepath, batch_id)
        return {"type": "notification", "rows": rows, "count": len(rows)}
    elif file_type == 'work_order':
        rows = parse_work_order_file(filepath, batch_id)
        return {"type": "work_order", "rows": rows, "count": len(rows)}
    elif file_type == 'bom':
        rows = parse_bom_file(filepath, batch_id)
        return {"type": "bom", "rows": rows, "count": len(rows)}
    else:
        return {"type": "unknown", "rows": [], "count": 0}